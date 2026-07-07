#!/usr/bin/env python3
"""
Construire un inventaire exploitable des donnees completes du projet McGill.

Sorties:
  - reports/mcgill_complete_inventory.csv
  - reports/mcgill_complete_inventory.md
"""

from __future__ import annotations

import csv
import re
from collections import Counter
from pathlib import Path
from zipfile import ZipFile

import openpyxl


ROOT = Path(__file__).resolve().parent
DATA_ROOT = ROOT / "Données completes"
REPORTS_DIR = ROOT / "reports"
CSV_OUT = REPORTS_DIR / "mcgill_complete_inventory.csv"
MD_OUT = REPORTS_DIR / "mcgill_complete_inventory.md"


def classify_path(path: Path) -> tuple[str, str]:
    rel = str(path.relative_to(ROOT))
    rel_lower = rel.lower()

    season = "other"
    for label in ["fall 2019", "winter 2019", "summer 2019", "fall 2021", "winter 2020", "summer 2020", "summer 2021"]:
        if label in rel_lower:
            season = label
            break

    source_type = "other"
    if "hobo" in rel_lower:
        source_type = "environment_hobo"
    elif "behavior" in rel_lower or "scan sampling" in rel_lower or "scan" in path.name.lower():
        source_type = "behavior_scan"
    elif "icetag" in rel_lower:
        source_type = "accelerometer_icetag"
    elif path.suffix.lower() == ".pdf":
        source_type = "article_or_pdf"
    elif path.suffix.lower() == ".docx":
        source_type = "documentation"

    if path.name == "Scan_Tot.xlsx":
        source_type = "behavior_aggregate"

    return season, source_type


def first_non_empty_rows(ws, n: int = 3) -> list[list[str]]:
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        values = ["" if v is None else str(v) for v in row]
        if any(v.strip() for v in values):
            rows.append(values[:15])
        if len(rows) >= n:
            break
    return rows


def summarize_xlsx(path: Path) -> dict:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = wb.sheetnames
    sample_sheet = sheets[0] if sheets else ""
    sample_rows = []
    if sample_sheet:
        sample_rows = first_non_empty_rows(wb[sample_sheet], n=3)
    wb.close()
    return {
        "sheet_count": len(sheets),
        "sheet_names_preview": " | ".join(sheets[:8]),
        "sample_sheet": sample_sheet,
        "sample_rows_preview": " || ".join(" | ".join(r) for r in sample_rows),
    }


def extract_docx_text(path: Path, max_chars: int = 2500) -> str:
    with ZipFile(path) as zf:
        xml = zf.read("word/document.xml").decode("utf-8")
    text = re.sub(r"<[^>]+>", "\n", xml)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()[:max_chars]


def build_inventory() -> list[dict]:
    rows: list[dict] = []
    for path in sorted(DATA_ROOT.rglob("*")):
        if not path.is_file():
            continue
        if path.name == ".DS_Store":
            continue

        season, source_type = classify_path(path)
        row = {
            "relative_path": str(path.relative_to(ROOT)),
            "filename": path.name,
            "extension": path.suffix.lower(),
            "season": season,
            "source_type": source_type,
            "sheet_count": "",
            "sheet_names_preview": "",
            "sample_sheet": "",
            "sample_rows_preview": "",
            "notes": "",
        }

        if path.suffix.lower() == ".xlsx":
            try:
                xlsx_info = summarize_xlsx(path)
                row.update(xlsx_info)
            except Exception as exc:
                row["notes"] = f"xlsx_read_error: {exc}"

        if path.suffix.lower() == ".docx":
            try:
                row["notes"] = extract_docx_text(path)
            except Exception as exc:
                row["notes"] = f"docx_read_error: {exc}"

        rows.append(row)

    return rows


def write_csv(rows: list[dict]) -> None:
    REPORTS_DIR.mkdir(exist_ok=True)
    fieldnames = [
        "relative_path",
        "filename",
        "extension",
        "season",
        "source_type",
        "sheet_count",
        "sheet_names_preview",
        "sample_sheet",
        "sample_rows_preview",
        "notes",
    ]
    with CSV_OUT.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_markdown(rows: list[dict]) -> str:
    total_files = len(rows)
    by_type = Counter(row["source_type"] for row in rows)
    by_season = Counter(row["season"] for row in rows)

    priority_lines = [
        "- Winter 2019: meilleur point de depart pour une analyse integree, car `cow_scan_long-format_7.12.19.xlsx` contient deja comportement + environnement + mesures IceTag synchrones.",
        "- Fall 2019: tres riche pour une integration multi-source, avec IceTag minute, HOBO et scans comportementaux, mais necessite davantage de fusion entre fichiers.",
        "- Summer 2019: riche egalement, avec scans, THI/HOBO et tableaux IceTag avant/apres/trajets, mais structure plus heterogene.",
        "- Fall 2021: semble surtout contenir des scans comportementaux; a confirmer si d'autres sources existent hors de ce dossier.",
        "- Documentation: le fichier `Étapes pour nettoyer les données.docx` explicite deja plusieurs regles de nettoyage utilisees pour les analyses de l'article.",
    ]

    cleaning_doc = next((row for row in rows if row["filename"] == "Étapes pour nettoyer les données.docx"), None)
    cleaning_excerpt = cleaning_doc["notes"][:1200] if cleaning_doc else "Non lu."

    lines = [
        "# Inventaire Complet Des Donnees McGill",
        "",
        "Date: 2026-04-16",
        "",
        "## Resume",
        "",
        f"- Nombre total de fichiers inventories: {total_files}",
        "- Cet inventaire couvre le dossier `Données completes` et vise a identifier les sources accelerometres, environnementales, comportementales et documentaires.",
        "- L'objectif est de preparer la phase d'integration multi-sources du projet McGill sans repartir de zero a chaque exploration.",
        "",
        "## Repartition Par Type De Source",
        "",
    ]

    for key, value in sorted(by_type.items()):
        lines.append(f"- {key}: {value}")

    lines.extend([
        "",
        "## Repartition Par Saison / Bloc D'etude",
        "",
    ])

    for key, value in sorted(by_season.items()):
        lines.append(f"- {key}: {value}")

    lines.extend([
        "",
        "## Priorites De Travail Recommandees",
        "",
    ])
    lines.extend(priority_lines)

    lines.extend([
        "",
        "## Extrait De La Documentation De Nettoyage",
        "",
        "Le document de nettoyage confirme deja plusieurs decisions de pretraitement importantes pour les analyses de l'article:",
        "",
        "```text",
        cleaning_excerpt,
        "```",
        "",
        "## Tables A Privilegier Des Maintenant",
        "",
        "- `Données completes/Données accelerometres/Winter 2019/cow_scan_long-format_7.12.19.xlsx`",
        "- `Données completes/Données accelerometres/Fall 2019/Icetag/IceTags/@IceTag_Compiled_Lameness.xlsx`",
        "- `Données completes/Données accelerometres/Fall 2019/HOBO data/.../*.xlsx`",
        "- `Données completes/Données accelerometres/Fall 2019/Behavioral scans - Amir's lameness project.xlsx`",
        "- `Données completes/Données accelerometres/Summer 2019/Behavioral scans - Catherine.xlsx`",
        "- `Données completes/Données accelerometres/Summer 2019/Hobo/THI per day of observation.xlsx`",
        "- `Données completes/Scan_Tot.xlsx`",
        "",
        "## Prochaine Etape Conseillee",
        "",
        "Construire une table d'integration cible par saison, en commencant par Winter 2019, car c'est la source la plus directement synchrone pour relier comportement, environnement et locomotion.",
    ])

    return "\n".join(lines) + "\n"


def main() -> None:
    rows = build_inventory()
    write_csv(rows)
    MD_OUT.write_text(build_markdown(rows), encoding="utf-8")
    print(f"CSV : {CSV_OUT}")
    print(f"MD  : {MD_OUT}")
    print(f"Fichiers inventories : {len(rows)}")


if __name__ == "__main__":
    main()
