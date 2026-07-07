#!/usr/bin/env python3
"""
Convertir les données IceTag McGill (Fall 2019) au format du pipeline de détection.

Source : @IceTag_Compiled_Lameness.xlsx (30 feuilles, 1/vache, 1 ligne/minute)
Sortie : mcgill_brut.csv (format identique à data/brut.csv, 1 ligne/15 min)

Transformation :
  - Chaque 15 minutes consécutives sont agrégées en 1 bin
  - Steps = somme sur le bin
  - Motion Index = somme sur le bin
  - Lying Time = somme des secondes couchées → format H:MM:SS
  - Standing Time = somme des secondes debout → format H:MM:SS
  - Transitions = somme des Lying Bouts
  - Transitions Up = Transitions // 2  (IceTag ne distingue pas Up/Down)
  - Transitions Down = Transitions - Transitions Up

Usage :
  python convert_mcgill_to_pipeline.py
  python convert_mcgill_to_pipeline.py --output chemin/vers/sortie.csv
"""

from __future__ import annotations

import argparse
import csv
import datetime
from collections import defaultdict
from pathlib import Path

import openpyxl


# === Chemins ===
SCRIPT_DIR = Path(__file__).resolve().parent
ICETAG_XLSX = (
    SCRIPT_DIR
    / "Données completes"
    / "Données accelerometres"
    / "Fall 2019"
    / "Icetag"
    / "IceTags"
    / "@IceTag_Compiled_Lameness.xlsx"
)
DEFAULT_OUTPUT = SCRIPT_DIR / "mcgill_brut.csv"

# === Labels SLS (scores de boiterie, Winter 2019) ===
# Source : Exercise Study - SLS Scores.xlsx
# Score = Edge + Rest + Shiftwt + Uneven (0-8)
# Les vaches avec SLS >= 2 aux DEUX évaluations sont classées "boiteuses persistantes"
SLS_LABELS = {
    # Cow: (SLS Baseline jan2019, SLS Midway mar2019, statut)
    821:  (2, 2, "boiteuse_persistante"),
    3444: (2, 4, "boiteuse_aggravee"),
    5874: (2, 1, "boiteuse_amelioree"),
    5857: (2, 1, "boiteuse_amelioree"),
    2081: (1, None, "legere"),
    2063: (1, 0, "legere_guerie"),
    2057: (1, 2, "legere_aggravee"),
    3437: (0, 1, "saine_degradee"),
    5875: (0, 1, "saine_degradee"),
    2078: (0, 0, "saine"),
    5879: (0, 0, "saine"),
    8500: (0, 1, "saine_degradee"),
}

# === Colonnes du fichier source ===
# Row: (Cow, Block, Treatment, Week, Date, Time, Motion Index,
#        Standing [t], Lying [t], Steps, Lying Bouts)
COL_COW = 0
COL_DATE = 4
COL_TIME = 5
COL_MI = 6
COL_STANDING = 7
COL_LYING = 8
COL_STEPS = 9
COL_LB = 10

BIN_MINUTES = 15


def time_to_seconds(t) -> int:
    """Convertir un datetime.time en secondes totales."""
    if t is None:
        return 0
    if isinstance(t, datetime.time):
        return t.hour * 3600 + t.minute * 60 + t.second
    if isinstance(t, datetime.timedelta):
        return int(t.total_seconds())
    return 0


def seconds_to_hms(total_seconds: int) -> str:
    """Convertir des secondes en format H:MM:SS (identique à brut.csv)."""
    h = total_seconds // 3600
    m = (total_seconds % 3600) // 60
    s = total_seconds % 60
    return f"{h}:{m:02d}:{s:02d}"


def make_bin_key(dt: datetime.datetime) -> datetime.datetime:
    """Arrondir un datetime au début du bin de 15 min."""
    minute_bin = (dt.minute // BIN_MINUTES) * BIN_MINUTES
    return dt.replace(minute=minute_bin, second=0, microsecond=0)


def process_sheet(ws, cow_id: str) -> list[dict]:
    """Traiter une feuille Excel (1 vache) → liste de bins 15 min."""

    # Collecter toutes les lignes minute par minute
    minutes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[COL_DATE]
        time_val = row[COL_TIME]

        if date_val is None or time_val is None:
            continue

        # Construire le datetime
        if isinstance(date_val, datetime.datetime):
            d = date_val.date()
        elif isinstance(date_val, datetime.date):
            d = date_val
        else:
            continue

        if isinstance(time_val, datetime.time):
            t = time_val
        elif isinstance(time_val, datetime.datetime):
            t = time_val.time()
        else:
            continue

        dt = datetime.datetime.combine(d, t)

        minutes.append({
            "dt": dt,
            "mi": int(row[COL_MI] or 0),
            "standing_s": time_to_seconds(row[COL_STANDING]),
            "lying_s": time_to_seconds(row[COL_LYING]),
            "steps": int(row[COL_STEPS] or 0),
            "lb": int(row[COL_LB] or 0),
        })

    # Trier par datetime
    minutes.sort(key=lambda x: x["dt"])

    # Agréger par bins de 15 min
    bins: dict[datetime.datetime, dict] = defaultdict(
        lambda: {"mi": 0, "standing_s": 0, "lying_s": 0, "steps": 0, "lb": 0, "count": 0}
    )
    for m in minutes:
        bk = make_bin_key(m["dt"])
        b = bins[bk]
        b["mi"] += m["mi"]
        b["standing_s"] += m["standing_s"]
        b["lying_s"] += m["lying_s"]
        b["steps"] += m["steps"]
        b["lb"] += m["lb"]
        b["count"] += 1

    # Convertir en liste de dicts au format brut.csv
    results = []
    for bin_start in sorted(bins.keys()):
        b = bins[bin_start]

        # Ne garder que les bins avec au moins 10 minutes de données
        # (pour éviter les bins quasi-vides en début/fin de journée)
        if b["count"] < 10:
            continue

        bin_end = bin_start + datetime.timedelta(minutes=BIN_MINUTES)
        transitions = b["lb"]
        transitions_up = transitions // 2
        transitions_down = transitions - transitions_up

        results.append({
            "Cow": cow_id,
            "Start": bin_start.strftime("%Y-%m-%d %H:%M:%S"),
            "End": bin_end.strftime("%Y-%m-%d %H:%M:%S"),
            "Steps": b["steps"],
            "Motion Index": b["mi"],
            "Lying Time": seconds_to_hms(b["lying_s"]),
            "Standing Time": seconds_to_hms(b["standing_s"]),
            "Transitions": transitions,
            "Transitions Up": transitions_up,
            "Transitions Down": transitions_down,
        })

    return results


def main():
    parser = argparse.ArgumentParser(
        description="Convertir les données IceTag McGill au format pipeline."
    )
    parser.add_argument(
        "--input", type=Path, default=ICETAG_XLSX,
        help="Fichier Excel source (défaut: @IceTag_Compiled_Lameness.xlsx)"
    )
    parser.add_argument(
        "--output", type=Path, default=DEFAULT_OUTPUT,
        help="Fichier CSV de sortie (défaut: mcgill_brut.csv)"
    )
    parser.add_argument(
        "--labels-csv", type=Path, default=None,
        help="Si fourni, exporte aussi un CSV avec les labels SLS par vache"
    )
    args = parser.parse_args()

    print(f"Lecture de {args.input.name}...")
    wb = openpyxl.load_workbook(str(args.input), read_only=True, data_only=True)

    all_rows = []
    cow_stats = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cow_id = sheet_name.strip()
        print(f"  Vache {cow_id}...", end=" ", flush=True)

        rows = process_sheet(ws, cow_id)
        all_rows.extend(rows)

        # Stats
        if rows:
            dates = set(r["Start"][:10] for r in rows)
            sls_info = SLS_LABELS.get(int(cow_id), (None, None, "pas_de_score"))
            cow_stats[cow_id] = {
                "bins": len(rows),
                "days": len(dates),
                "first": rows[0]["Start"][:10],
                "last": rows[-1]["Start"][:10],
                "sls_baseline": sls_info[0],
                "sls_midway": sls_info[1],
                "statut": sls_info[2],
            }
            print(f"{len(rows)} bins, {len(dates)} jours, SLS: {sls_info[2]}")
        else:
            print("AUCUNE donnée!")

    wb.close()

    # Trier par vache puis par date
    all_rows.sort(key=lambda r: (r["Cow"], r["Start"]))

    # Écriture du CSV
    fieldnames = [
        "Cow", "Start", "End", "Steps", "Motion Index",
        "Lying Time", "Standing Time", "Transitions",
        "Transitions Up", "Transitions Down",
    ]
    with open(args.output, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_rows)

    print(f"\n{'=' * 60}")
    print(f"RÉSULTAT : {args.output}")
    print(f"  {len(all_rows)} lignes (bins de 15 min)")
    print(f"  {len(cow_stats)} vaches")
    print(f"  Période : {min(r['Start'] for r in all_rows)[:10]} → "
          f"{max(r['Start'] for r in all_rows)[:10]}")

    # Export labels SLS
    labels_path = args.labels_csv or args.output.parent / "mcgill_sls_labels.csv"
    with open(labels_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Cow", "SLS_Baseline_Jan2019", "SLS_Midway_Mar2019",
                         "Statut", "Bins_15min", "Jours", "Debut", "Fin"])
        for cow_id in sorted(cow_stats.keys(), key=int):
            s = cow_stats[cow_id]
            writer.writerow([
                cow_id, s["sls_baseline"], s["sls_midway"], s["statut"],
                s["bins"], s["days"], s["first"], s["last"],
            ])
    print(f"  Labels SLS : {labels_path}")

    # Résumé des groupes cliniques
    print(f"\n{'=' * 60}")
    print("GROUPES CLINIQUES (vaches avec score SLS disponible) :")
    groups = defaultdict(list)
    for cow_id, stats in cow_stats.items():
        groups[stats["statut"]].append(cow_id)
    for statut in sorted(groups.keys()):
        cows = groups[statut]
        print(f"  {statut:<25} : {len(cows)} vaches → {', '.join(sorted(cows, key=int))}")

    print(f"\n✅ Prêt pour le pipeline ! Commande :")
    print(f"   cp {args.output} /Users/alioubarry/PROJECT/data/mcgill_brut.csv")


if __name__ == "__main__":
    main()
